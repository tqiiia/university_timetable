import datetime
from openpyxl import load_workbook

letters = {'A': 0, 'B': 1,
               'C': 2, 'D': 3,
               'E': 4, 'F': 5,
               'G': 6, 'H': 7,
               'I': 8, 'J': 9,
               'K': 10, 'L': 11,
               'M': 12, 'N': 13,
               'O': 14, 'P': 15,
               'Q': 16, 'R': 17}
weekdays = {'Понедельник': 0,
            'Вторник': 1,
            'Среда': 2,
            'Четверг': 3,
            'Пятница': 4,
            'Суббота': 5,
            'Воскресенье': 6}


async def schedule_week(group, even=0):
    all_lessons = ''
    return_data = ''
    with open('groups.txt', 'r') as groups:
        for g in groups:
            if g.split('\t')[0] == group:
                group = g
                book = load_workbook(group.split('\t')[1], read_only=True)
                sheet = book[group.split('\t')[2]]
                for i in range(int(group.split('\t')[3][1:]) + 1, 120):
                    lesson = sheet[group.split('\t')[3][0] + f'{i}'].value
                    if lesson:
                        date = sheet['B' + f'{i}'].value
                        j = 1
                        while not date:
                            date = sheet['B' + f'{i - j}'].value
                            j += 1
                        time = sheet['D' + f'{i}'].value
                        if not time:
                            time = 'чётн.' + str(sheet['D' + f'{i - 1}'].value)
                        auditorium = sheet[list(letters.keys())[list(letters.values()).index(
                            list(letters.keys()).index(group.split('\t')[3][0]) + 1)] + f'{i}'].value
                        if lesson == 'ДСП':
                            auditorium = ''
                            time = '09.00-20.30'
                        if not auditorium and lesson != 'ДСП':
                            auditorium = sheet[list(letters.keys())[list(letters.values()).index(
                                list(letters.keys()).index(group.split('\t')[3][0]) + 3)] + f'{i}'].value
                            if not auditorium:
                                pass
                            else:
                                all_lessons += str(date) + '\t' + str(time) + '\t' + str(lesson) + '\t' + str(auditorium) + '\n'
                        else:
                            all_lessons += str(date) + '\t' + str(time) + '\t' + str(lesson) + '\t' + str(auditorium) + '\n'
    all_lessons = all_lessons.rstrip()

    print(all_lessons)

    even_add = 1 if not (int(datetime.date(2023, 9, 1).isocalendar()[1]) % 2 == 0) else 0
    for i in range(0, 6):
        for lesson in all_lessons.split('\n'):
            if lesson.split('\t')[0].rstrip() == list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))]:
                if return_data.find(str(list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))])) == -1:
                    return_data += f'<b>{str(list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))])}</b>\n'
                if (int(datetime.datetime.today().isocalendar()[1]) + even_add + even) % 2 != 0:
                    return_data += str(lesson.replace('чётн.', '* ').split('\t', 1)[1]) + '\n'
                else:
                    if lesson.split('\t')[1].find('чётн.'):
                        return_data += str(lesson.split('\t', 1)[1]) + '\n'
    print(return_data)

    n = 0
    if (int(datetime.datetime.today().isocalendar()[1]) + even_add + even) % 2 != 0:
        print('чётная', even_add, even)
        for data in return_data.split('\n'):
            if data.find('*') == 0 and (
                    return_data.split('\n')[n].split('\t')[0] == data.split('\t')[0].replace('* ', '')) == 0 \
                    and return_data.split('\n')[n - 1].find('*') != 0:
                text = return_data.split('\n')[n - 1]
                if len(text.split('\t')) > 1:
                    return_data = return_data.replace(f"{text}", "\n")
                    n += 1
            n += 1
    else:
        print('нечётная', even_add, even)

    return_data = return_data.replace('\n\n', '').replace('* ', '')
    print(return_data)

    if return_data:
        return return_data


async def schedule_day(group, offset=0):
    all_lessons = ''
    daily_lessons = ''
    return_data = ''
    date_now = datetime.datetime.today().weekday()
    with open('groups.txt', 'r') as groups:
        for g in groups:
            if g.split('\t')[0] == group:
                group = g
                book = load_workbook(group.split('\t')[1], read_only=True)
                sheet = book[group.split('\t')[2]]
                for i in range(int(group.split('\t')[3][1:]) + 1, 120):
                    lesson = sheet[group.split('\t')[3][0] + f'{i}'].value
                    if lesson:
                        date = sheet['B' + f'{i}'].value
                        j = 1
                        while not date:
                            date = sheet['B' + f'{i - j}'].value
                            j += 1
                        time = sheet['D' + f'{i}'].value
                        if not time:
                            time = 'чётн.' + str(sheet['D' + f'{i - 1}'].value)
                        auditorium = sheet[list(letters.keys())[list(letters.values()).index(
                            list(letters.keys()).index(group.split('\t')[3][0]) + 1)] + f'{i}'].value
                        if lesson == 'ДСП':
                            auditorium = ''
                            time = '09.00-20.30'
                        if not auditorium and lesson != 'ДСП':
                            auditorium = sheet[list(letters.keys())[list(letters.values()).index(
                                list(letters.keys()).index(group.split('\t')[3][0]) + 3)] + f'{i}'].value
                            if not auditorium:
                                pass
                            else:
                                all_lessons += str(date) + '\t' + str(time) + '\t' + str(lesson) + '\t' + str(
                                    auditorium) + '\n'
                        else:
                            all_lessons += str(date) + '\t' + str(time) + '\t' + str(lesson) + '\t' + str(
                                auditorium) + '\n'

    for lesson in all_lessons.split('\n'):
        date = str(lesson.split('\t')[0])
        if offset == 1 and datetime.datetime.today().weekday() + offset == 7:
            offset = -6
        if date.rstrip() == str(
                list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(date_now + offset))]):
            daily_lessons += lesson + '\n'

    print(daily_lessons)

    even_add = 1 if not (int(datetime.date(2023, 9, 5).isocalendar()[1]) % 2 == 0) else 0
    even_add = even_add + 1 if offset == -6 else even_add
    # even_add = even_add if not (int(datetime.datetime.today().weekday()) == 6) else even_add - 6
    for i in range(0, 6):
        for lesson in daily_lessons.split('\n'):
            if lesson.split('\t')[0].rstrip() == list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))]:
                if return_data.find(str(
                        list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))])) == -1:
                    return_data += f'<b>{str(list(weekdays.keys())[list(weekdays.values()).index(list(weekdays.values()).index(i))])}</b>\n'
                if (int(datetime.datetime.today().isocalendar()[1]) + even_add) % 2 != 0:
                    return_data += str(lesson.replace('чётн.', '* ').split('\t', 1)[1]) + '\n'
                else:
                    if lesson.split('\t')[1].find('чётн.'):
                        return_data += str(lesson.split('\t', 1)[1]) + '\n'
    print(return_data)

    n = 0
    if (int(datetime.datetime.today().isocalendar()[1]) + even_add) % 2 != 0:
        print('чётная', even_add)
        for data in return_data.split('\n'):
            if data.find('*') == 0 and (
                    return_data.split('\n')[n].split('\t')[0] == data.split('\t')[0].replace('* ', '')) == 0 \
                    and return_data.split('\n')[n - 1].find('*') != 0:
                text = return_data.split('\n')[n - 1]
                if len(text.split('\t')) > 1:
                    return_data = return_data.replace(f"{text}", "\n")
                    n += 1
            n += 1
    else:
        print('нечётная', even_add)

    return_data = return_data.replace('\n\n', '').replace('* ', '')

    print(return_data)

    if not return_data:
        return_data = 'Занятий на этот день нет'
    return return_data
