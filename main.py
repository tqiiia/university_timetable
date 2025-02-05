import re
import glob
from openpyxl import load_workbook

file_names = str(glob.glob('*.xlsx'))

pattern_time = re.compile(r'\b\d\d.\d\d-\d\d.\d\d\b')
pattern_discipline = re.compile(r'\b\d.\d.')
pattern_group = re.compile(r'\b\d\d.\d\d.\d\d \b')

letters = {'A': 0,  'B': 1,
           'C': 2,  'D': 3,
           'E': 4,  'F': 5,
           'G': 6,  'H': 7,
           'I': 8,  'J': 9,
           'K': 10, 'L': 11,
           'M': 12, 'N': 13,
           'O': 14, 'P': 15,
           'Q': 16, 'R': 17}

#   Использовать для заполнения файла groups.txt
with open('groups.txt', 'w') as g:
    g.close()
for file_name in file_names.rsplit(','):
    file_name = str(file_name).replace(' ', '').replace("'", "").replace("[", "").replace("]", "").rstrip()
    print(file_name)
    book = load_workbook(file_name, read_only=True)
    for sheet in book:
        for letter in letters:
            for i in range(1, 150):
                #   num group
                if type(sheet[letter + str(i)].value) == str and re.match(pattern_group, sheet[letter + str(i)].value):
                    with open('groups.txt', 'a') as g:
                        g.write(str(sheet[letter + str(i)].value).split(' ')
                                [int(str(sheet[letter + str(i)].value).split(' ').index('гр.')) + 1]
                                + '\t' + str(file_name)
                                + '\t' + str(sheet.title)
                                + '\t' + str(sheet[letter + str(i)].coordinate) + '\n')

# with open('groups.txt', 'r') as groups:
#     for group in groups:
#         book = load_workbook(group.split('\t')[1], read_only=True)
#         sheet = book[group.split('\t')[2]]
#         # print(list(letters.keys())[list(letters.values()).index(list(letters.keys()).index(group.split('\t')[3][0]) + 1)])
#         print('\n\n')
#         print(sheet[group.split('\t')[3]].value)
#         for i in range(int(group.split('\t')[3][1:]) + 1, 120):
#             lesson = sheet[group.split('\t')[3][0] + f'{i}'].value
#             if lesson:
#                 date = sheet['B' + f'{i}'].value
#                 j = 1
#                 while not date:
#                     date = sheet['B' + f'{i-j}'].value
#                     j += 1
#                 time = sheet['D' + f'{i}'].value
#                 if not time:
#                     time = 'чётн.' + str(sheet['D' + f'{i-1}'].value)
#                 auditorium = sheet[list(letters.keys())[list(letters.values()).index(list(letters.keys()).index(group.split('\t')[3][0]) + 1)] + f'{i}'].value
#                 if not auditorium and lesson != 'ДСП':
#                     auditorium = sheet[list(letters.keys())[list(letters.values()).index(
#                         list(letters.keys()).index(group.split('\t')[3][0]) + 3)] + f'{i}'].value
#                     if not auditorium:
#                         auditorium = sheet[list(letters.keys())[list(letters.values()).index(
#                             list(letters.keys()).index(group.split('\t')[3][0]) + 5)] + f'{i}'].value
#                     else:
#                         print(date, time, lesson, auditorium)
#                 else:
#                     print(date, time, lesson, auditorium)
